# parsed_info.py

import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from constants import (
    EXCEL_OUTPUT, DESIRED_COLUMNS,
    CENSUS_MAPPING, CARGO_FLAGS
)

def format_date_fmsca(val):
    if not val:
        return ""
    try:
        if " " in val:
            dt = datetime.strptime(val.split()[0], "%Y%m%d")
        else:
            dt = datetime.strptime(val, "%Y%m%d")
        return dt.strftime("%Y-%m-%d")
    except:
        return val

def get_full_address(street, city, state, zip_code, country="US"):
    parts = [street, city, f"{state} {zip_code}", country]
    return ", ".join(filter(None, parts)).strip(", ")

def get_cargo_transported(company):
    cargos = []
    for flag in CARGO_FLAGS:
        if company.get(flag) == "X":
            name = flag.replace("crgo_", "").replace("_", " ").title()
            cargos.append(name)
    return ", ".join(cargos) if cargos else "[нет]"

def process_companies(from_excel=None):


    print(f"Загружаем DOT из: {from_excel}")
    wb = load_workbook(from_excel)
    ws = wb.active
    dots = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]

    print(f"Найдено DOT: {len(dots)}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Result"
    out_ws.append(DESIRED_COLUMNS)

    for dot in dots:
        print(f"\n--- DOT {dot} ---")

        # Census
        census_url = f"https://data.transportation.gov/resource/az4n-8mr2.json?dot_number={dot}"
        try:
            r = requests.get(census_url, timeout=12)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"Census ошибка: {e}")
            out_ws.append([dot] + ["Ошибка"] * (len(DESIRED_COLUMNS) - 1))
            continue

        if not data:
            out_ws.append([dot] + ["Не найдено"] * (len(DESIRED_COLUMNS) - 1))
            continue

        company = data[0]

        row_data = {}
        for api_key, col_name in CENSUS_MAPPING.items():
            val = company.get(api_key, "")
            if col_name == "Added":
                val = format_date_fmsca(val)
            row_data[col_name] = val

        row_data["Cargo Transported"] = get_cargo_transported(company)

        # MC
        mc = ""
        for p in ["docket1prefix", "docket2prefix", "docket3prefix"]:
            pref = company.get(p, "")
            num = company.get(p.replace("prefix", ""), "")
            if pref and num:
                mc = f"{pref}{num}"
                break
        row_data["MC/MX/FF"] = mc

        # Адреса — в одну ячейку каждый
        row_data["Physical Address"] = get_full_address(
            company.get("phy_street"), company.get("phy_city"),
            company.get("phy_state"), company.get("phy_zip")
        )
        row_data["Mailing Address"] = get_full_address(
            company.get("carrier_mailing_street"), company.get("carrier_mailing_city"),
            company.get("carrier_mailing_state"), company.get("carrier_mailing_zip")
        )

        # Страховки
        insurance_rows = []
        if mc:
            insur_url = f"https://data.transportation.gov/resource/qh9u-swkp.json?docket_number={mc}&$limit=50"
            try:
                r_ins = requests.get(insur_url, timeout=12)
                r_ins.raise_for_status()
                policies = r_ins.json()

                if policies:
                    def parse_dt(d):
                        try: return datetime.strptime(d, "%m/%d/%Y")
                        except: return datetime(1900, 1, 1)

                    policies.sort(key=lambda x: parse_dt(x.get("effective_date", "")), reverse=True)

                    for pol in policies:
                        ins_dict = {
                            "TypeInsurance": pol.get("ins_form_code", "[нет]"),
                            "Carrier": pol.get("name_company", "[нет]"),
                            "Policy/Surety": pol.get("policy_no", "[нет]"),
                            "Coverage": pol.get("max_cov_amount", "[нет]"),
                            "Effective Date": pol.get("effective_date", "[нет]"),
                            "Cancellation Date": pol.get("cancl_effective_date", "[не отменена]")
                        }
                        insurance_rows.append(ins_dict)
            except Exception as e:
                print(f"Страховка ошибка: {e}")

        # Формируем строку
        final_row = []
        for col in DESIRED_COLUMNS:
            value = row_data.get(col, "[нет]")

            if col in ["TypeInsurance", "Carrier", "Policy/Surety", "Coverage", "Effective Date", "Cancellation Date"]:
                if insurance_rows:
                    value = insurance_rows[0].get(col, "[нет]")
            elif col.startswith("Insurance2"):
                idx = 1
                base = col.replace("Insurance2 ", "").strip()
                field_map = {
                    "Type": "TypeInsurance",
                    "Carrier": "Carrier",
                    "Policy/Surety": "Policy/Surety",
                    "Coverage": "Coverage",
                    "Effective Date": "Effective Date",
                    "Cancellation Date": "Cancellation Date"
                }
                real_field = field_map.get(base, base)
                value = insurance_rows[idx].get(real_field, "[нет]") if len(insurance_rows) > idx else "[нет]"
            elif col.startswith("Insurance3"):
                idx = 2
                base = col.replace("Insurance3 ", "").strip()
                field_map = {
                    "Type": "TypeInsurance",
                    "Carrier": "Carrier",
                    "Policy/Surety": "Policy/Surety",
                    "Coverage": "Coverage",
                    "Effective Date": "Effective Date",
                    "Cancellation Date": "Cancellation Date"
                }
                real_field = field_map.get(base, base)
                value = insurance_rows[idx].get(real_field, "[нет]") if len(insurance_rows) > idx else "[нет]"

            final_row.append(value)

        out_ws.append(final_row)
        print(f"DOT {dot} — OK (страховок: {len(insurance_rows)})")

    out_wb.save(EXCEL_OUTPUT)
    print(f"\nГотово! Сохранено: {EXCEL_OUTPUT}")
    return f"Обработка завершена. Результат в {EXCEL_OUTPUT}"