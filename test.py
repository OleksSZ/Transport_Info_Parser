import requests
import json
from openpyxl import load_workbook

# === Загружаем Excel и читаем DOT из первой колонки ===
excel_file = r"C:\Users\User\Desktop\Broker_automat\Dot.xlsx"   # ← имя твоего файла
wb = load_workbook(excel_file)
ws = wb.active

dot_numbers = []
for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # начиная со 2-й строки
    if row[0]:
        dot_numbers.append(str(row[0]))

print(f"Найдено DOT номеров: {len(dot_numbers)}\n")

# === Обрабатываем каждый DOT ===
for DOT_NUMBER in dot_numbers:
    print("=" * 80)
    print(f"\n>>> Обработка DOT {DOT_NUMBER}\n")

    # --- Census ---
    census_url = f"https://data.transportation.gov/resource/az4n-8mr2.json?dot_number={DOT_NUMBER}"
    try:
        census_resp = requests.get(census_url, timeout=10)
        census_resp.raise_for_status()
        census_data = census_resp.json()

        if not census_data:
            print(f"Компания с DOT {DOT_NUMBER} не найдена в Census.")
            continue

        company = census_data[0]
        print(f"\n=== Данные Census по DOT {DOT_NUMBER} ===")
        for key, value in sorted(company.items()):
            if not value:
                value = "[пусто]"
            print(f"{key: <35} : {value}")

        # Берём MC номер
        mc_number = company.get("docket1", "")
        mc_prefix = company.get("docket1prefix", "")
        if mc_number and mc_prefix.upper() == "MC":
            mc_full = f"{mc_prefix}{mc_number}"
            print(f"\nMC номер для поиска страховок: {mc_full}")
        else:
            print("\nMC номер не найден в Census.")
            continue

    except Exception as e:
        print(f"Ошибка Census: {e}")
        continue

    # --- Страховки ---
    insur_url = f"https://data.transportation.gov/resource/qh9u-swkp.json?docket_number={mc_full}&$limit=100000"
    try:
        insur_resp = requests.get(insur_url, timeout=10)
        insur_resp.raise_for_status()
        insur_data = insur_resp.json()

        if not insur_data:
            print("\nСтраховки не найдены.")
            continue

        print(f"\n=== Страховки по {mc_full} ===")
        print(f"Всего найдено: {len(insur_data)} записей\n")
        for i, pol in enumerate(insur_data, 1):
            print(f"Полис {i}:")
            for key, value in pol.items():
                if not value:
                    value = "[пусто]"
                print(f"  {key: <25} : {value}")
            print("─" * 60)

    except Exception as e:
        print(f"Ошибка страховок: {e}")
        continue
